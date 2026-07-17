#!/usr/bin/env python3
"""Resolve the ACTUAL allowed values of an Amazon template's dropdown fields.

The `accepted_values` text in Data Definitions is prose, not an enforceable list.
The real constraint lives in the `Template` sheet's data validations, which point
(directly, via named range, or via INDIRECT) at ranges on the `Dropdown Lists` /
`Valid Values` sheets. This resolves them into `column -> allowed[]` so the agent
can pick legal values and `write_values.py` / `validate_output.py` can enforce them.

How each validation `formula1` is resolved:
  - literal  `"A,B,C"`        -> those values
  - name     `record_action`  -> defined name -> range values
  - INDIRECT(...&"suffix")     -> product_type prefix + the literal suffix embedded
                                  in the formula = a defined name -> range values
  - `Sheet!$A$1:$A$9`          -> read that range
Anything we cannot resolve is emitted with "resolved": false (enum_unresolved) and
a reason — never silently treated as "anything allowed".

Usage:
  python3 resolve_valid_values.py TEMPLATE.xlsm [--out valid_values.json]
"""
import argparse
import json
import re
import sys

import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter

from parse_template import load_settings


def read_range(wb, ref):
    """Read non-empty cell values from an A1-style range ref like 'Sheet'!$A$4:$A$9."""
    m = re.match(r"^'?([^'!]+)'?!\$?([A-Za-z]+)\$?(\d+)(?::\$?([A-Za-z]+)\$?(\d+))?$", ref)
    if not m:
        return None
    sheet, c1, r1, c2, r2 = m.group(1), m.group(2), int(m.group(3)), m.group(4), m.group(5)
    c2 = c2 or c1
    r2 = int(r2) if r2 else r1
    if sheet not in wb.sheetnames:
        return None
    ws = wb[sheet]
    out = []
    for col in range(column_index_from_string(c1), column_index_from_string(c2) + 1):
        for row in range(r1, r2 + 1):
            v = ws.cell(row=row, column=col).value
            if v not in (None, ""):
                out.append(str(v).strip())
    return out


def lookup_name(wb, name):
    """Return the range ref string for a defined name, or None."""
    try:
        dn = wb.defined_names[name]
    except (KeyError, TypeError):
        return None
    return dn.value if dn else None


def indirect_target_name(formula, product_type):
    """Compute the defined-name an INDIRECT(...) formula resolves to.

    Mirrors the template's own logic:
      IF(ISNUMBER(VALUE(LEFT(pt,1))),"_","") & SUBSTITUTE(SUBSTITUTE(pt,"-","_")," ","") & "<suffix>"
    The <suffix> is the last quoted literal in the formula.
    """
    literals = re.findall(r'"([^"]*)"', formula)
    if not literals:
        return None
    suffix = literals[-1]
    pt = (product_type or "").replace("-", "_").replace(" ", "")
    prefix = "_" if pt[:1].isdigit() else ""
    return f"{prefix}{pt}{suffix}"


def resolve_formula(wb, formula, product_type):
    """Return (values|None, source, reason). values=None means unresolved."""
    f = str(formula or "").strip()
    if not f:
        return None, "none", "empty formula"
    # literal comma list, optionally quoted
    if f.startswith('"') and f.endswith('"'):
        inner = f[1:-1]
        return [v.strip() for v in inner.split(",") if v.strip()], "literal", ""
    if f.upper().startswith("INDIRECT"):
        name = indirect_target_name(f, product_type)
        if not name:
            return None, "indirect", "could not parse suffix"
        ref = lookup_name(wb, name)
        if not ref:
            return None, "indirect", f"defined name not found: {name}"
        vals = read_range(wb, ref)
        return (vals, "indirect", "" ) if vals is not None else (None, "indirect", f"range unreadable: {ref}")
    if "!" in f:  # direct sheet range
        vals = read_range(wb, f)
        return (vals, "range", "") if vals is not None else (None, "range", f"range unreadable: {f}")
    # bare token -> defined name
    ref = lookup_name(wb, f)
    if ref:
        vals = read_range(wb, ref)
        return (vals, "named-range", "") if vals is not None else (None, "named-range", f"range unreadable: {ref}")
    return None, "unknown", f"unrecognized formula: {f[:40]}"


def sqref_columns(sqref, data_row):
    """Yield 1-based column indexes covered by a validation sqref that touch data rows."""
    cols = set()
    for part in str(sqref).split():
        m = re.match(r"^\$?([A-Za-z]+)\$?(\d+)(?::\$?([A-Za-z]+)\$?(\d+))?$", part)
        if not m:
            continue
        c1, r1, c2, r2 = m.group(1), int(m.group(2)), m.group(3), m.group(4)
        c2 = c2 or c1
        r2 = int(r2) if r2 else r1
        if r2 < data_row:  # header-only validation (e.g. row 4)
            continue
        for c in range(column_index_from_string(c1), column_index_from_string(c2) + 1):
            cols.add(c)
    return cols


def build(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Template"]
    settings = load_settings(ws)
    data_row, attr_row = settings["data_row"], settings["attribute_row"]
    product_type = settings.get("product_type")
    attrs = [c.value for c in next(ws.iter_rows(min_row=attr_row, max_row=attr_row))]

    columns = {}
    for dv in ws.data_validations.dataValidation:
        if dv.type != "list":
            continue
        vals, source, reason = resolve_formula(wb, dv.formula1, product_type)
        for col in sqref_columns(dv.sqref, data_row):
            attr = attrs[col - 1] if col - 1 < len(attrs) else None
            columns[col] = {
                "column": col,
                "column_letter": get_column_letter(col),
                "attribute": (str(attr).strip() if attr else None),
                "resolved": vals is not None,
                "allowed": vals,
                "source": source,
                "reason": reason,
            }
    return {"template": path, "product_type": product_type,
            "enum_columns": len(columns),
            "resolved": sum(1 for c in columns.values() if c["resolved"]),
            "columns": columns}


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("template")
    ap.add_argument("--out", default="valid_values.json")
    args = ap.parse_args()
    manifest = build(args.template)
    with open(args.out, "w", encoding="utf-8") as fh:
        json.dump(manifest, fh, ensure_ascii=False, indent=2)
    print(f"Product type: {manifest['product_type']}")
    print(f"Enum (dropdown) columns: {manifest['enum_columns']}  "
          f"resolved: {manifest['resolved']}  -> {args.out}")
    unresolved = [c for c in manifest["columns"].values() if not c["resolved"]]
    if unresolved:
        print(f"\nUnresolved ({len(unresolved)}) — enforced as enum_unresolved (warn only):")
        for c in unresolved[:10]:
            print(f"  col {c['column']} {c['attribute']}: {c['reason']}")
    # show a couple resolved samples
    print("\nSample resolved:")
    for c in list(manifest["columns"].values())[:4]:
        if c["resolved"]:
            print(f"  {c['attribute']}: {c['allowed'][:6]}{' …' if len(c['allowed'])>6 else ''}")


if __name__ == "__main__":
    main()
