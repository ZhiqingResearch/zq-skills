#!/usr/bin/env python3
"""Write resolved field values into an Amazon flat-file template, with enforcement.

Reads a values.json (v2 schema) and writes each value into the `Template` sheet
from the data row down, applying these gates so the output is trustworthy:

  - **Field policy** (field_policy.py):
      * compliance fields (country of origin, battery, dangerous goods, FCC, Prop
        65…) are NEVER written from a guess — if `inferred` or lacking a source,
        they are left blank (needs_user_input).
      * seller_owned fields (SKU, price, condition, fulfillment, warranty…) are
        never written from a web/Keepa source — they must come from user intake.
  - **Enum enforcement** (valid_values.json from resolve_valid_values.py): a value
    for a resolved dropdown column that isn't an allowed value is a HARD ERROR — the
    file is NOT saved. Casing is auto-corrected to the canonical allowed value.
  - **Identity gate**: a row whose `identity_confidence` is "low" is skipped (not
    written) unless --force.
  - Inferred cells are highlighted; the template's built-in example data is cleared
    (unless --keep-examples); macros, other sheets, and dropdown validation are
    preserved; output keeps the original extension.

values.json v2:
{
  "template": "NOTEBOOK_COMPUTER.xlsm",
  "output": "…",                 // optional; default <name>.filled<ext>
  "clear_examples": true,        // optional; default true
  "rows": [
    {
      "upc": "889842188837",
      "identity_confidence": "high",     // high | medium | low
      "identity_evidence": ["UPC exact match on brand site", "MPN matches Keepa"],
      "values": {
        "<attribute>": {
          "value": "Dell", "source": "keepa", "source_url": "https://…",
          "evidence": "Brand: Dell", "confidence": "high",
          "inferred": false, "status": "filled"   // filled|inferred|needs_user_input|blocked
        }
      }
    }
  ]
}

Usage:
  python3 write_values.py values.json [--valid-values valid_values.json]
                                      [--color FFE0B2] [--keep-examples] [--force]
"""
import argparse
import json
import os
import re
import sys

import openpyxl
from openpyxl.styles import PatternFill

from field_policy import classify

WEB_SOURCES = {"web_search", "web", "google", "keepa", "bing", "search"}


def load_settings(ws):
    raw = str(ws.cell(row=1, column=1).value or "")
    def grab(name, default):
        m = re.search(rf"{name}=(\d+)", raw)
        return int(m.group(1)) if m else default
    return grab("attributeRow", 5), grab("dataRow", 8)


def clear_region(ws, from_row, to_row, max_col):
    cleared, sample = 0, []
    for r in range(from_row, to_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            if cell.value is not None:
                if len(sample) < 8:
                    sample.append(f"R{r}C{c}={str(cell.value)[:18]}")
                cell.value = None
                cleared += 1
    return cleared, sample


def norm(s):
    return str(s).strip().lower() if s is not None else ""


def load_valid_values(path):
    """col(int) -> {canon: {normalized_value: canonical_value}, raw: [...]} for resolved enums."""
    if not path:
        return {}
    with open(path, encoding="utf-8") as fh:
        data = json.load(fh)
    out = {}
    for col_str, info in (data.get("columns") or {}).items():
        if info.get("resolved") and info.get("allowed"):
            canon = {norm(v): v for v in info["allowed"]}
            out[int(info["column"])] = {"canon": canon, "raw": info["allowed"]}
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("values")
    ap.add_argument("--valid-values", help="valid_values.json for enum enforcement")
    ap.add_argument("--color", default="FFE0B2",
                    help="hex fill for inferred cells (default light orange)")
    ap.add_argument("--keep-examples", action="store_true",
                    help="do NOT clear the template's example data")
    ap.add_argument("--force", action="store_true",
                    help="write rows even if identity_confidence is low")
    args = ap.parse_args()

    with open(args.values, encoding="utf-8") as fh:
        spec = json.load(fh)

    template = spec["template"]
    base, ext = os.path.splitext(template)
    output = spec.get("output") or f"{base}.filled{ext}"
    is_macro = template.lower().endswith(".xlsm")
    clear_examples = spec.get("clear_examples", True) and not args.keep_examples
    valid = load_valid_values(args.valid_values)

    wb = openpyxl.load_workbook(template, keep_vba=is_macro, data_only=False)
    ws = wb["Template"]
    attr_row, data_row = load_settings(ws)
    max_col = ws.max_column
    col_of = {}
    for cell in next(ws.iter_rows(min_row=attr_row, max_row=attr_row)):
        if cell.value is not None:
            col_of[str(cell.value).strip()] = cell.column

    # ---- Validation pass (no writes): decide what to write, collect gates ----
    plan = []                 # (row, col, canonical_value, inferred)
    enum_errors = []          # HARD: illegal enum values -> no file produced
    warnings = []
    skipped = {"compliance": 0, "seller_web": 0, "needs_user_input": 0,
               "identity_row": 0, "unmatched": 0, "enum_ok_corrected": 0}

    for i, row in enumerate(spec.get("rows", [])):
        r = data_row + i
        upc = row.get("upc", "?")
        ident = norm(row.get("identity_confidence"))
        if not ident:
            warnings.append(f"row {r} (upc {upc}): no identity_confidence provided")
        row_blocked = ident == "low" and not args.force
        if row_blocked:
            skipped["identity_row"] += 1
            warnings.append(f"row {r} (upc {upc}): identity_confidence low -> row NOT written (use --force)")
            continue

        for attr, info in (row.get("values") or {}).items():
            col = col_of.get(attr.strip())
            if not col:
                skipped["unmatched"] += 1
                continue
            if not isinstance(info, dict):
                info = {"value": info}
            value = info.get("value")
            if value in (None, ""):
                continue
            status = norm(info.get("status"))
            if status in ("needs_user_input", "blocked"):
                skipped["needs_user_input"] += 1
                continue
            inferred = bool(info.get("inferred"))
            source = norm(info.get("source"))
            has_source = bool(info.get("source_url") or info.get("evidence"))
            policy = classify(attr)

            if policy == "compliance" and (inferred or not has_source):
                skipped["compliance"] += 1
                warnings.append(f"compliance '{attr[:44]}' not firmly sourced -> left blank (needs_user_input)")
                continue
            if policy == "seller_owned" and source in WEB_SOURCES:
                skipped["seller_web"] += 1
                warnings.append(f"seller field '{attr[:44]}' has web source '{source}' -> left blank (provide via intake)")
                continue

            v = valid.get(col)
            if v is not None:
                key = norm(value)
                if key not in v["canon"]:
                    enum_errors.append((r, upc, attr, value, v["raw"]))
                    continue
                canonical = v["canon"][key]
                if canonical != value:
                    skipped["enum_ok_corrected"] += 1
                value = canonical

            plan.append((r, col, value, inferred))

    if enum_errors:
        print(f"✗ ABORTED — {len(enum_errors)} illegal dropdown value(s); no file written:\n", file=sys.stderr)
        for r, upc, attr, value, allowed in enum_errors[:20]:
            shown = ", ".join(map(str, allowed[:8])) + (" …" if len(allowed) > 8 else "")
            print(f"  row {r} (upc {upc}) {attr[:40]}\n    got: {value!r}\n    allowed: {shown}", file=sys.stderr)
        sys.exit(1)

    # ---- Write pass ----
    cleared, sample = 0, []
    if clear_examples:
        last = max(ws.max_row, data_row + len(spec.get("rows", [])) - 1)
        cleared, sample = clear_region(ws, attr_row + 1, last, max_col)

    fill = PatternFill(start_color=args.color, end_color=args.color, fill_type="solid")
    written = inferred_n = 0
    for r, col, value, inferred in plan:
        cell = ws.cell(row=r, column=col, value=value)
        written += 1
        if inferred:
            cell.fill = fill
            inferred_n += 1

    wb.save(output)

    # ---- Report ----
    print(f"Saved: {output}")
    if clear_examples:
        print(f"Example data cleared: {cleared} cell(s)" + (f"  e.g. {', '.join(sample[:5])}" if sample else ""))
    print(f"Cells written: {written}  (inferred/highlighted: {inferred_n})")
    if valid:
        print(f"Enum enforcement: ON ({len(valid)} resolved dropdown columns)"
              + (f", {skipped['enum_ok_corrected']} casing-corrected" if skipped["enum_ok_corrected"] else ""))
    else:
        print("Enum enforcement: OFF (pass --valid-values valid_values.json to enable)")
    gates = {k: v for k, v in skipped.items() if v and k != "enum_ok_corrected"}
    if gates:
        print("Skipped by gate:", gates)
    if warnings:
        print(f"\nWarnings ({len(warnings)}):", file=sys.stderr)
        for w in warnings[:25]:
            print("  - " + w, file=sys.stderr)


if __name__ == "__main__":
    main()
