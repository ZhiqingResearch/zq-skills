#!/usr/bin/env python3
"""Parse an Amazon flat-file listing template (.xlsm/.xlsx) into a field manifest.

Amazon templates encode their layout in the `Template` sheet's A1 settings blob
(labelRow / attributeRow / dataRow) and describe every field — including its
filling rules and whether it is required — in the `Data Definitions` sheet.

This script merges the two so an agent knows, per column:
  - the technical attribute name (matches Template attribute row)
  - the human label + group
  - the Accepted Values text (the FILLING RULE)
  - the required level: Required / Conditionally Required / Recommended / Optional
  - the 1-based column index in the Template sheet

Usage:
  python3 parse_template.py TEMPLATE.xlsm [--out fields.json] [--required-only]

Prints a short summary to stdout and writes the full manifest as JSON.
"""
import argparse
import json
import re
import sys

import openpyxl


def load_settings(ws):
    """Read labelRow/attributeRow/dataRow from the Template!A1 settings string."""
    raw = ws.cell(row=1, column=1).value or ""
    def grab(name, default):
        m = re.search(rf"{name}=(\d+)", str(raw))
        return int(m.group(1)) if m else default
    return {
        "label_row": grab("labelRow", 4),
        "attribute_row": grab("attributeRow", 5),
        "data_row": grab("dataRow", 8),
    }


def row_values(ws, row_idx):
    return [c.value for c in next(ws.iter_rows(min_row=row_idx, max_row=row_idx))]


def load_definitions(wb):
    """Map field attribute name -> {group, local_label, accepted_values, example, required}."""
    dd = wb["Data Definitions"]
    rows = list(dd.iter_rows(values_only=True))
    # Locate the header row (contains "Field Name" and "Required?").
    header_idx = None
    for i, r in enumerate(rows):
        cells = [str(c) for c in r if c is not None]
        if any("Field Name" in c for c in cells) and any("Required" in c for c in cells):
            header_idx = i
            break
    if header_idx is None:
        raise SystemExit("Could not find the Data Definitions header row.")

    header = [str(c).strip() if c is not None else "" for c in rows[header_idx]]
    def col(name):
        for j, h in enumerate(header):
            if h.lower().startswith(name.lower()):
                return j
        return None
    ci = {k: col(v) for k, v in {
        "group": "Group Name", "field": "Field Name", "label": "Local Label",
        "accepted": "Accepted Values", "example": "Example", "required": "Required",
    }.items()}

    defs = {}
    group = ""
    for r in rows[header_idx + 1:]:
        g = r[ci["group"]] if ci["group"] is not None else None
        if g:
            group = str(g).strip()
        field = r[ci["field"]] if ci["field"] is not None else None
        if not field:
            continue
        key = str(field).strip()
        defs[key] = {
            "group": group,
            "local_label": _s(r, ci["label"]),
            "accepted_values": _s(r, ci["accepted"]),
            "example": _s(r, ci["example"]),
            "required": _s(r, ci["required"]) or "Optional",
        }
    return defs


def _s(row, idx):
    if idx is None:
        return ""
    v = row[idx]
    return "" if v is None else str(v).strip()


def build_manifest(path):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb["Template"]
    settings = load_settings(ws)
    attrs = row_values(ws, settings["attribute_row"])
    labels = row_values(ws, settings["label_row"])
    defs = load_definitions(wb)

    fields = []
    for idx, attr in enumerate(attrs):
        if attr is None:
            continue
        attr = str(attr).strip()
        d = defs.get(attr, {})
        fields.append({
            "column": idx + 1,  # 1-based, matches openpyxl
            "attribute": attr,
            "label": (str(labels[idx]).strip() if idx < len(labels) and labels[idx] else d.get("local_label", "")),
            "group": d.get("group", ""),
            "required": d.get("required", "Unknown"),
            "accepted_values": d.get("accepted_values", ""),
            "example": d.get("example", ""),
        })
    return {"template": path, **settings, "sheet": "Template",
            "field_count": len(fields), "fields": fields}


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("template")
    ap.add_argument("--out", default="fields.json")
    ap.add_argument("--required-only", action="store_true",
                    help="Keep only Required / Conditionally Required fields")
    args = ap.parse_args()

    manifest = build_manifest(args.template)
    if args.required_only:
        manifest["fields"] = [f for f in manifest["fields"]
                              if "Required" in f["required"]]

    with open(args.out, "w", encoding="utf-8") as fh:
        json.dump(manifest, fh, ensure_ascii=False, indent=2)

    from collections import Counter
    dist = Counter(f["required"] for f in manifest["fields"])
    print(f"Template: {manifest['template']}")
    print(f"Layout: label_row={manifest['label_row']} "
          f"attribute_row={manifest['attribute_row']} data_row={manifest['data_row']}")
    print(f"Fields written: {len(manifest['fields'])}  ->  {args.out}")
    print("Required-level distribution:", dict(dist))
    req = [f for f in manifest["fields"] if f["required"] == "Required"]
    print(f"\nStrictly Required ({len(req)}):")
    for f in req:
        print(f"  col {f['column']:>3}  {f['label']}")


if __name__ == "__main__":
    main()
