#!/usr/bin/env python3
"""Validate a filled Amazon template before it is handed to the user.

Independently re-checks the OUTPUT file (not just the write log) and exits non-zero
on any hard ERROR. Severity:
  ERROR (blocks) — enum value not allowed; a dropdown/sheet/macro lost vs the
                   original; file won't re-open.
  WARN           — a Required product field is blank; UPC check-digit fails; a
                   value/unit pair is half-filled; residual example data remains.
  INFO           — a Required compliance/seller field is blank (expected: needs
                   user input).

Usage:
  python3 validate_output.py FILLED.xlsm --template ORIGINAL.xlsm \
      --fields fields.json --valid-values valid_values.json [--json report.json]
"""
import argparse
import json
import re
import sys
import zipfile

import openpyxl

from field_policy import classify
from parse_template import load_definitions, load_settings


def gtin_check_ok(code):
    s = re.sub(r"\D", "", str(code))
    if len(s) not in (12, 13, 14):
        return None  # not a GTIN we check
    digits = [int(c) for c in s]
    check = digits[-1]
    body = digits[:-1][::-1]
    total = sum(d * (3 if i % 2 == 0 else 1) for i, d in enumerate(body))
    return (10 - total % 10) % 10 == check


def struct_counts(path):
    z = zipfile.ZipFile(path)
    names = z.namelist()
    dv = sum(z.read(n).count(b"<dataValidation ")
             for n in names if n.startswith("xl/worksheets") and n.endswith(".xml"))
    return {"vba": "xl/vbaProject.bin" in names, "data_validations": dv}


def data_rows(ws, data_row, max_col):
    last = data_row - 1
    for r in range(data_row, ws.max_row + 1):
        if any(ws.cell(row=r, column=c).value not in (None, "") for c in range(1, max_col + 1)):
            last = r
    return range(data_row, last + 1)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("filled")
    ap.add_argument("--template", help="original template, for structure comparison")
    ap.add_argument("--fields", help="fields.json (for Required checks)")
    ap.add_argument("--valid-values", help="valid_values.json (for enum checks)")
    ap.add_argument("--json", help="write the machine-readable report here")
    args = ap.parse_args()

    issues = []
    def add(sev, msg):
        issues.append({"severity": sev, "message": msg})

    # Re-open (corruption check)
    try:
        wb = openpyxl.load_workbook(args.filled, data_only=True)
    except Exception as e:  # noqa: BLE001
        add("ERROR", f"output does not re-open: {e}")
        _emit(issues, args.json)
        sys.exit(1)
    ws = wb["Template"]
    settings = load_settings(ws)
    data_row = settings["data_row"]
    attrs = [c.value for c in next(ws.iter_rows(min_row=settings["attribute_row"],
                                                max_row=settings["attribute_row"]))]
    max_col = ws.max_column
    rows = list(data_rows(ws, data_row, max_col))
    col_attr = {i + 1: (str(a).strip() if a else None) for i, a in enumerate(attrs)}

    # Structure preserved vs original
    if args.template:
        o, f = struct_counts(args.template), struct_counts(args.filled)
        if o["vba"] and not f["vba"]:
            add("ERROR", "macros (vbaProject.bin) lost vs original")
        if f["data_validations"] < o["data_validations"]:
            add("ERROR", f"dropdown validations dropped: {o['data_validations']} -> {f['data_validations']}")
        ob = set(openpyxl.load_workbook(args.template, read_only=True).sheetnames)
        if ob - set(wb.sheetnames):
            add("ERROR", f"sheets missing vs original: {sorted(ob - set(wb.sheetnames))}")

    # Enum membership (independent re-check)
    if args.valid_values:
        with open(args.valid_values, encoding="utf-8") as fh:
            vv = json.load(fh)
        allowed = {int(c["column"]): {str(x).strip().lower() for x in c["allowed"]}
                   for c in (vv.get("columns") or {}).values() if c.get("resolved") and c.get("allowed")}
        for r in rows:
            for col, aset in allowed.items():
                v = ws.cell(row=r, column=col).value
                if v not in (None, "") and str(v).strip().lower() not in aset:
                    add("ERROR", f"row {r} col {col} ({col_attr.get(col)}): illegal enum value {v!r}")

    # Required-field presence (policy-aware) + upload-readiness verdict
    needs_user = 0
    if args.fields:
        with open(args.fields, encoding="utf-8") as fh:
            fields = json.load(fh)["fields"]
        req = [f for f in fields if f["required"] == "Required"]
        for r in rows:
            for f in req:
                if ws.cell(row=r, column=f["column"]).value in (None, ""):
                    pol = classify(f["attribute"])
                    sev = "WARN" if pol == "product_attribute" else "INFO"
                    if pol in ("compliance", "seller_owned"):
                        needs_user += 1
                    add(sev, f"row {r}: Required '{f['label']}' blank ({pol})")
        if needs_user:
            add("INFO", f"UPLOAD READINESS: attributes-only — {needs_user} required "
                        f"compliance/seller field(s) need user input before upload")
        else:
            add("INFO", "UPLOAD READINESS: all required fields present")

    # UPC / product-id check digit
    pid_col = next((c for c, a in col_attr.items() if a and "product_id_value" in a), None)
    if pid_col:
        for r in rows:
            v = ws.cell(row=r, column=pid_col).value
            if v not in (None, "") and gtin_check_ok(v) is False:
                add("WARN", f"row {r}: product id {v!r} fails GTIN check digit")

    # Value/unit pairing (…#N.value with a sibling …#N.unit)
    unit_cols = {a: c for c, a in col_attr.items() if a and a.endswith(".unit")}
    for r in rows:
        for c, a in col_attr.items():
            if a and a.endswith(".value"):
                sib = a[:-6] + ".unit"
                if sib in unit_cols:
                    has_v = ws.cell(row=r, column=c).value not in (None, "")
                    has_u = ws.cell(row=r, column=unit_cols[sib]).value not in (None, "")
                    if has_v != has_u:
                        add("WARN", f"row {r}: '{a}' value/unit pair half-filled")

    # Residual example data (values still matching Data Definitions Example column)
    if args.template:
        try:
            defs = load_definitions(openpyxl.load_workbook(args.template, data_only=True, read_only=True))
            ex = {str(d["example"]).strip().lower() for d in defs.values()
                  if d.get("example") and len(str(d["example"]).strip()) > 2}
            for r in rows:
                for c, a in col_attr.items():
                    v = ws.cell(row=r, column=c).value
                    if v not in (None, "") and str(v).strip().lower() in ex:
                        add("WARN", f"row {r} col {c} ({a}): value {v!r} matches template example — possible residual")
        except Exception:  # noqa: BLE001
            pass

    _emit(issues, args.json)
    errors = sum(1 for i in issues if i["severity"] == "ERROR")
    sys.exit(1 if errors else 0)


def _emit(issues, json_path):
    from collections import Counter
    counts = Counter(i["severity"] for i in issues)
    if json_path:
        with open(json_path, "w", encoding="utf-8") as fh:
            json.dump({"summary": dict(counts), "issues": issues}, fh, ensure_ascii=False, indent=2)
    verdict = "FAIL" if counts.get("ERROR") else ("PASS (with warnings)" if counts.get("WARN") else "PASS")
    print(f"Validation: {verdict}  "
          f"[ERROR={counts.get('ERROR', 0)} WARN={counts.get('WARN', 0)} INFO={counts.get('INFO', 0)}]")
    order = {"ERROR": 0, "WARN": 1, "INFO": 2}
    for i in sorted(issues, key=lambda x: order[x["severity"]])[:40]:
        print(f"  [{i['severity']}] {i['message']}")


if __name__ == "__main__":
    main()
